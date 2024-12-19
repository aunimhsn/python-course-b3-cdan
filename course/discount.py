from openpyxl import load_workbook

wb = load_workbook('./data/transactions.xlsx')
ws = wb['sheet_1']

for i in range(2, ws.max_row + 1):
    current_price:str = ws[f'C{i}'].value
    parsed_price = float(current_price.replace(',', '.').replace('€', ''))
    discounted_price = parsed_price * 0.9
    ws[f'D{i}'] = f'{discounted_price}€'

wb.save('./data/transactions_updated.xlsx')